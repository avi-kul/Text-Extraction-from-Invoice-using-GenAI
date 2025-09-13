import re
import os
import pandas as pd
from PIL import Image
from io import BytesIO
import streamlit as st
from dotenv import load_dotenv
from openpyxl import load_workbook
import google.generativeai as genai


# Loading the Environmental variables
load_dotenv() 

# GenAI instance 
# genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))
genai.configure(api_key='')

print(genai.configure)  # Ensure this prints your actual API key


# model = genai.GenerativeModel('gemini-pro-vision')
model = genai.GenerativeModel('gemini-1.5-flash')
# model = genai.GenerativeModel('chat-bison-001')


def get_gemini_response(input,image):
    if not image:
        return "No image provided."
    response = model.generate_content([input_prompt, image[0]])
    return response.text if response else "No response from model."


def input_image_details(uploaded_file):
    if uploaded_file is not None:
        
        # read the file into bytes 
        bytes_data = uploaded_file.getvalue()

        image_parts = [
            {
                "mime_type" : uploaded_file.type, # get the mime type of the uploaded file
                "data" : bytes_data
            }
        ]
        return image_parts
    return []  # return empty list if no file
    


# Convert the extracted text to a dictionary and clean the keys
def extract_to_dict(response_text):
    # Split the response by lines
    lines = response_text.split('\n')
    
    # Create a dictionary to store the extracted details
    invoice_details = {}
    
    for line in lines:
        # Split the line by ':' and add to the dictionary
        if ':' in line:
            key, value = line.split(':', 1)
            # Clean the key using regex to remove unwanted characters like '-' and strip spaces
            cleaned_key = re.sub(r'^[-\s]+|[-\s]+$', '', key).strip()
            invoice_details[cleaned_key] = value.strip()
    
    return invoice_details


# Function to save invoice data to an Excel file
def save_invoice_to_excel(invoice_data):
    """
    Save the invoice details to the 'Invoices.xlsx' file.
    If the file exists, append the data. If not, create a new file.
    :param invoice_data: A dictionary containing invoice details
    """
    # Define the Excel file name
    excel_file = "Invoice_Details.xlsx"

    # Create a DataFrame for the new invoice entry
    new_data = pd.DataFrame([invoice_data])  # Convert dictionary to DataFrame

    # Check if the Excel file exists
    if os.path.exists(excel_file):
        try:
            # Load the existing workbook and get the sheet
            workbook = load_workbook(excel_file)
            sheet = workbook.active  # Assuming we are appending to the active sheet

            # Find the last row with data
            start_row = sheet.max_row
            
            # Append the new data to the Excel file
            with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists='overlay') as writer:
                new_data.to_excel(writer, index=False, header=False, startrow=start_row, sheet_name=sheet.title)
        
        except Exception as e:
            print(f"Error reading or appending to Excel file: {str(e)}")
    else:
        # If the file doesn't exist, create a new file and save the data
        new_data.to_excel(excel_file, index=False, sheet_name="Invoices")

    print(f"Invoice details saved to {excel_file}.")

# Function to provide download link for the Excel file
def provide_download_link(file_name):
    with open(file_name, 'rb') as f:
        data = f.read()
    
    return data

##### Streamlit setup #####


# st.set_page_config(page_title='Invoice Details Extraction')

st.write(" ")
st.header('Get details from the Invoice in one click...!')
st.write(" ")

# input = st.text_input('Input Prompt : ',key='input')
uploaded_file = st.file_uploader("Please upload an image of Invoice...",type=['jpg','jpeg','png'])
image = ''

if uploaded_file is not None:
    image = Image.open(uploaded_file)
    st.image(image,caption='Image successfully uploaded...!',use_column_width=True)


submit = st.button("Click here to generate the Details")

st.write("_______________________________________________")



# input_prompt = """
# You are an expert in understanding invoices. We will upload an image as invoice and you will have to answer 
# any questions based on the uploaded invoice image 
# """

input_prompt = """
You are an expert in understanding invoices. We will upload an image as invoice and you will have to answer 
any questions based on the uploaded invoice image.
We are specifically looking to extract the following details:
- Invoice number
- Consignee name
- Vendor name
- Total amount
"""

# if submit button is clicked

# if submit:
#     image_data = input_image_details(uploaded_file)
#     response = get_gemini_response(input_prompt,image_data,input)
#     st.subheader("The Response is")
#     st.write(response)
#     print(response)
#     print(type(response[0]))

    
# If submit button is clicked
if submit:
    image_data = input_image_details(uploaded_file)
    response = get_gemini_response(input_prompt, image_data)
    
    # Display the response
    st.subheader("Extracted Data from the Invoice is :- ")
    st.write(response)
    st.write(" ")
    st.write("_"*50)
    
    # Convert response to dictionary
    invoice_data = extract_to_dict(response)
    # st.subheader("Extracted Invoice Data as Dictionary")
    # st.write(invoice_data)
    
    # Append the data to Excel
    excel_file_name = 'Invoice_Details.xlsx'
    save_invoice_to_excel(invoice_data)
    
    # Provide download link for the updated Excel file
    excel_data = provide_download_link(excel_file_name)
    
    st.download_button(
        label="Download Updated Excel",
        data=excel_data,
        file_name=excel_file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


